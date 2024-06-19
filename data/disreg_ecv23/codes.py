import csv
import os
import shutil


from glob import glob
from collections import defaultdict

import openpyxl

CODE_DIR = "codes"

shutil.rmtree(CODE_DIR, ignore_errors=True)
os.mkdir(CODE_DIR)

maps = defaultdict(list)
code_map = {}

for file in glob("*.xlsx"):
    wb = openpyxl.load_workbook(file)
    for sheet in wb.worksheets:
        if sheet.title == "Diseño":
            rows = sheet.iter_rows(values_only=True)

            # skip two first rows
            next(rows)
            next(rows)

            # build dict map
            for row in rows:
                f1, f2 = row[:2]

                if f1 == "*** TOTAL ***":
                    break

                if f1 and f2:
                    code_map[f1] = f2
        elif sheet.title.startswith("Tablas"):
            map = None
            in_key = None
            rows = sheet.iter_rows(values_only=True)
            for row in rows:
                f1, f2 = row[:2]
                if not in_key and isinstance(f1, str):
                    map = f1
                    in_key = True
                    next(rows)
                elif in_key:
                    if f1 is not None:
                        maps[map].append((f1, f2))
                    else:
                        in_key = False

# Write CSVs
for code, map in code_map.items():
    print([code, map])
    with open(f"{CODE_DIR}/{code}.csv", "w") as codefile:
        code_writer = csv.writer(codefile)
        code_writer.writerow(["id", "description"])
        for key, val in maps[map]:
            code_writer.writerow([key, val])
