import os
import csv
import shutil

from glob import glob
from collections import defaultdict

import duckdb
import openpyxl

CODE_DIR = "data/disreg_ecv23/codes"
DB_NAME = "db/ecv.duckdb"

# Build the codes csvs
shutil.rmtree(CODE_DIR, ignore_errors=True)
os.mkdir(CODE_DIR)

maps = defaultdict(list)
code_map = {}

for file in glob("data/disreg_ecv23/*.xlsx"):
    wb = openpyxl.load_workbook(file)
    wb_name = os.path.splitext(os.path.basename(file))[0]
    for sheet in wb.worksheets:
        if sheet.title == "Diseño":
            rows = sheet.iter_rows(values_only=True)

            # skip two first rows
            next(rows)
            next(rows)

            # build dict map
            for row in rows:
                f1, f2 = row[:2]

                if f1 == "*** TOTAL ***":
                    break

                if f1 and f2:
                    code_map[f1] = f"{wb_name}_{f2}"
        elif sheet.title.startswith("Tablas"):
            map = None
            in_key = None
            rows = sheet.iter_rows(values_only=True)
            for row in rows:
                f1, f2 = row[:2]
                if not in_key and isinstance(f1, str):
                    map = f"{wb_name}_{f1}"
                    in_key = True
                    next(rows)
                elif in_key:
                    if f1 is not None:
                        maps[map].append((f1, f2))
                    else:
                        in_key = False

# Write CSVs
for map_name, map in maps.items():
    path = f"{CODE_DIR}/{map_name}.csv"
    if os.path.exists(path):
        raise Exception(f"ERROR: code already exists: {path}")
    with open(path, "w") as codefile:
        code_writer = csv.writer(codefile)
        code_writer.writerow(["id", "desc"])
        for key, val in map:
            code_writer.writerow([key, val])

# Build the database
try:
    os.remove(DB_NAME)
except OSError:
    pass

con = duckdb.connect(database=DB_NAME)

for csv_path in glob("data/**/*.csv", recursive=True):
    for path in glob(csv_path):
        tbl = os.path.splitext(os.path.basename(path))[0]
        con.execute(f"CREATE TABLE {tbl} AS SELECT * FROM read_csv('{path}')")

# create views (translates from codes to maps)
for code, map in code_map.items():
    con.execute(f"CREATE VIEW {code} AS SELECT * FROM {map};")
